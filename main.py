import shutil
from openpyxl import Workbook, load_workbook
import datetime
import random
import numpy as np
from dateutil import relativedelta
import collections

src = "./Budget Dataset Modified.xlsx"
dest = "./Output/Backup.xlsx"
shutil.copyfile(src, dest)

data = load_workbook("./Output/Backup.xlsx")
result_wb = Workbook()

data_ws = data["Project Details"]
project_ids = data_ws["A"][1:]
project_ids = [i.value for i in project_ids]
project_budgets = data_ws["L"][1:]
project_budgets = [i.value for i in project_budgets]

# Budget
if True:
    result_ws = result_wb.active
    result_ws.title = "Budget"

    # Headers
    result_ws.append(["Project ID", "Start Date", "Duration (Months)", "Budgeted Cost per Month", "Overall Budget", "Projected Income"])

    # date_style = openpyxl.styles.NamedStyle(name="date_style", number_format="dd/mm/yyyy")
    i = 0
    for row in result_ws.iter_rows(min_row=2, max_row=len(project_ids) + 1, max_col=6):
        # Project IDs
        row[0].value = project_ids[i]

        # Start Date
        # row[1].style = date_style
        # start = datetime.datetime.strptime("01/01/2020", "%d/%m/%Y").date()
        # start date will be in between 2 years before (min duration is 2 years) and 6 months before (need 6-24 months of data)
        start = datetime.date.today() - datetime.timedelta(days=365*2)
        end = datetime.date.today() - datetime.timedelta(days=30*6)
        delta = end - start
        r = random.randrange(delta.days)
        new = start + datetime.timedelta(days=r)
        row[1].value = new

        # Duration (Months)
        row[2].value = random.randrange(24, 60)

        # Cost per Month
        row[3].style = "Currency"
        row[3].value = project_budgets[i] / row[2].value

        # Total Cost
        row[4].style = "Currency"
        row[4].value = project_budgets[i]

        # Contract Pay
        row[5].style = "Currency"
        row[5].value = project_budgets[i] * random.randint(1020, 1080)/1000

        i += 1

# Categories
if True:
    result_wb.create_sheet("Categories")
    result_ws = result_wb["Categories"]

    #Headers
    result_ws.append(["Project ID", "Category", "Category Budget"])

    categories = [
        'Direct Labour',
        'Supplied Labour',
        'Sub-contractor',
        'Other Materials',
        'Small Tools & Safety Item',
        'Other Consumable',
        'Transportation',
        'Repair & Maintenance',
        'Site Office Expense',
        'Food, Refreshment & Entertainment',
        'Travelling & Vehicles',
        'Main Steel Materials',
        'Stainless Steel Materials',
        'Aluminium Materials',
        'Equipment',
        'Transportation',
        'Supervision',
        'Insurance'
        ]

    result_ws = result_wb["Budget"]
    budgets = result_ws["E"][1:]
    budgets = [i.value for i in budgets]
    result_ws = result_wb["Categories"]
    i2 = 0
    i4 = 1
    for i in project_ids:
        n, k = budgets[i2], len(categories)
        vals = np.random.default_rng().dirichlet(np.ones(k), size=1)
        k_nums = [round(v) for v in vals[0]*n]
        i3 = 0
        for category in categories:
            i4 += 1
            result_ws.append([i, category, k_nums[i3]])
            result_ws[f"C{i4}"].style = "Currency"
            i3 += 1
        i2 += 1

# Cash Outflow
if True:
    result_wb.create_sheet("Cash Outflow")
    result_ws = result_wb["Cash Outflow"]

    # Headers
    result_ws.append(["Project ID", "Date", "Category", "Actual Category Monthly Cost"])

    start_dates = result_wb["Budget"]["B"][1:]
    start_dates = [i.value for i in start_dates]
    duration = result_wb["Budget"]["C"][1:]
    duration = [i.value for i in duration]
    budgets_by_cat = result_wb["Categories"]["C"][1:]
    budgets_by_cat = [i.value for i in budgets_by_cat]

    i2 = 0
    i5 = 1
    for i in project_ids:
        the_date = start_dates[i2]
        the_date += relativedelta.relativedelta(months=1)
        while(datetime.date.today() - the_date > datetime.timedelta(days=0)):
            i4 = 0
            for category in categories:
                i5 += 1
                result_ws.append([i, the_date, category, budgets_by_cat[i4]/duration[i2]*(random.randint(900, 1100)/1000)])
                result_ws[f"D{i5}"].style = "Currency"
                i4 += 1
            the_date += relativedelta.relativedelta(months=1)
        i2 += 1

# Reports
if True:
    result_wb.copy_worksheet(result_wb["Cash Outflow"])
    result_ws = result_wb["Cash Outflow Copy"]
    result_ws.delete_cols(3, 2)

    values = []
    for row in result_ws.iter_rows(min_row=2):
        if [row[0].value, row[1].value] in values:
            pass
        else:
            values.append([row[0].value, row[1].value])

    result_wb.remove(result_ws)
    result_wb.create_sheet("Reports")
    result_ws = result_wb["Reports"]
    result_ws.append(["Project ID", "Date", "Completion", "Incurred Cost", "BCWS"])

    for value in values:
        result_ws.append(value)

    journey = result_ws["A"][1:]
    journey = [i.value for i in journey]

    journey = collections.Counter(journey)
    journey = journey.values()
    journey = list(journey)

    secret_tool_for_later = []
    l = []
    i2 = 0
    for i in duration:
        # i = Planned duration of project (Duration (Months)), journey[i2] = Months passed by
        secret_tool_for_later.append(journey[i2])
        random_variance = random.randint(80, 120)/100
        percentage = journey[i2]/i*random_variance
        accumulative = 0
        base_percentage = percentage/journey[i2]
        for i3 in range(journey[i2]):
            # For some epic sauce
            random_variance_2 = random.randint(80,120)/100
            base_percentage *= random_variance_2

            accumulative += base_percentage
            # Make sure it does not go above 100%
            if accumulative > 1:
                accumulative = 1
            l.append(accumulative)
        i2 += 1

    bruh = 2
    for value in l:
        result_ws[f"C{bruh}"] = value
        bruh += 1

    for cell in result_ws["C"]:
        cell.style = "Percent"

    # Incurred Cost oh my goodness gracious
    result_ws = result_wb["Cash Outflow"]
    count = 0
    the_things_i_want = {}
    for row in result_ws.iter_rows(min_row=2):
        the_pog_champ = f"{row[0].value}-{row[1].value}"
        if the_pog_champ not in the_things_i_want:
            the_things_i_want[the_pog_champ] = row[3].value
        else:
            the_things_i_want[the_pog_champ] += row[3].value

    l = []
    l2 = []
    current_project = ""
    for k in the_things_i_want:
        if current_project != k[:11]:
            current_project = k[:11]
            l.append(l2)
            l2 = []
        l2.append(the_things_i_want[k])
    l.append(l2)
    
    l3 = []
    for i in l:
        l3.append(list(np.cumsum(i)))
    
    l4 = []
    for i in l3:
        for i2 in i:
            l4.append(i2)
    
    result_ws = result_wb["Reports"]
    awesomesauce = result_ws["D"][1:]
    i = 0
    for cell in awesomesauce:
        cell.value = l4[i]
        cell.style = "Currency"
        i += 1

    # BCWS
    monthly_costs = result_wb["Budget"]["D"][1:]
    monthly_costs = [i.value for i in monthly_costs]
    l = []
    i3 = 0
    for i in secret_tool_for_later:
        for i2 in range(i):
            l.append((i2 + 1) * monthly_costs[i3])
        i3 += 1
    awesomesauce = result_ws["E"][1:]
    i = 0
    for cell in awesomesauce:
        cell.value = l[i]
        cell.style = "Currency"
        i += 1

# Cash Inflow
if True:
    result_wb.copy_worksheet(result_wb["Reports"])
    result_ws = result_wb["Reports Copy"]
    values = []
    for row in result_ws.iter_rows(min_row=2):
        if row[0].value not in [i[0] for i in values]:
            values.append([row[0].value, row[1].value, row[2].value])
        # Im so proud of this
        elif row[2].value >= 0.2 and [i[2] for i in values if i[0] == row[0].value and i[2] >= 0.2] == []:
            values.append([row[0].value, row[1].value, row[2].value])
        elif row[2].value >= 0.4 and [i[2] for i in values if i[0] == row[0].value and i[2] >= 0.4] == []:
            values.append([row[0].value, row[1].value, row[2].value])
        elif row[2].value >= 0.8 and [i[2] for i in values if i[0] == row[0].value and i[2] >= 0.8] == []:
            values.append([row[0].value, row[1].value, row[2].value])
        else:
            pass

    contract_pays = result_wb["Budget"]["F"][1:]
    contract_pays = [i.value for i in contract_pays]

    d = {}
    i2 = 0
    for i in project_ids:
        d[i] = contract_pays[i2]
        i2 += 1
        
    for i in values:
        i.pop()
        i.append(d[i[0]]/5)

    result_wb.remove(result_ws)
    result_wb.create_sheet("Cash Inflow")
    result_ws = result_wb["Cash Inflow"]

    # Headers
    result_ws.append(["Project ID", "Date", "Income"])

    for i in values:
        result_ws.append(i)

    for cell in result_ws["C"]:
        cell.style = "Currency"


# Project Details
if True:
    result_wb.create_sheet("Project Details", 0)
    result_ws = result_wb["Project Details"]

    l = []
    for row in data_ws.iter_rows():
        l2 = []
        for cell in row:
            l2.append(cell.value)
        l.append(l2)
    for i in l:
        result_ws.append(i)
    for cell in result_ws["L"]:
        cell.style = "Currency"

# Change status
if True:
    statuses_to_replace_with = ["Planning", "Planned", "Execution", "Finishing", "Finished"]
    result_ws = result_wb["Project Details"]

    statuses = result_ws["D"][1:]

    for i in statuses:
        random_swag_number = random.randrange(5)
        i.value = statuses_to_replace_with[random_swag_number]

# Regions
if True:
    regions = [
    "North-East",
    "North-East",
    "Central",
    "West",
    "West",
    "West",
    "Central",
    "Central",
    "West",
    "Central",
    "North",
    "East",
    "West",
    "West",
    "West",
    "Central",
    "Central",
    "North-East",
    "West",
    "West",
    "Central",
    "North",
    "North",
    "Central",
    "Central",
    "Central",
    "Central",
    "North-East",
    "North-East",
    "Central",
    "Central",
    "Central",
    "East",
    "East",
    "East"
    ]
    result_ws = result_wb["Project Details"]
    result_ws.insert_cols(4)

    region_col = result_ws["D"]
    region_col[0].value = "Regions"
    region_col = region_col[1:]
    i = 0
    for cell in region_col:
        cell.value = regions[i]
        i += 1

data.close()
result_wb.save("./Output/Result.xlsx")